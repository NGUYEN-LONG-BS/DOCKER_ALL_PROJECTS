from django.utils.timezone import now
from django.http import HttpResponse

from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.pagination import PageNumberPagination

from .models_LA import LA_INVENTORY_CATEGORIES
from .models_LA import LA_CLIENT_CATEGORIES

from .serializers_LA import LAClientCategoriesSerializer

import openpyxl
import re
from openpyxl.utils import get_column_letter
from datetime import timedelta

DATABASE_NAME = 'tala'
    
# ==============================================================================
# INVENTORY CATEGORIES
# ==============================================================================

# Import bulk data
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bulk_data_LA_INVENTORY_CATEGORIES(request):
    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if 'import-data' not in wb.sheetnames:
            return Response({'error': 'Không tìm thấy sheet "import-data" trong file Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['import-data']
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Sheet "import-data" không có dữ liệu.'}, status=status.HTTP_400_BAD_REQUEST)
        header = rows[0]
        header_map = {col: idx for idx, col in enumerate(header)}
        required_fields = ['id_nhan_vien', 'xoa_sua', 'ma_hang', 'ten_hang', 'dvt', 'sl_ton_dau_ky', 'don_gia_ton_dau_ky', 'ma_kho_luu_tru']
        for field in required_fields:
            if field not in header_map:
                return Response({'error': f'Thiếu cột bắt buộc: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra mã hàng trùng trước khi import
        for idx, row in enumerate(rows[1:], start=2):
            ma_hang = row[header_map['ma_hang']]
            # xoa_sua = row[header_map['xoa_sua']]
            if ma_hang and LA_INVENTORY_CATEGORIES.objects.using(DATABASE_NAME).filter(ma_hang=ma_hang, xoa_sua="new").exists():
                return Response({'error': f'Mã hàng "{ma_hang}" ở dòng {idx} đã tồn tại, vui lòng kiểm tra lại.'}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            # Kiểm tra dòng trống hoặc thiếu mã hàng
            if not row or not row[header_map['ma_hang']]:
                errors.append(f'Dòng {idx}: Cột "ma_hang" không được rỗng.')
                continue
            # Kiểm tra cột xoa_sua không được rỗng
            if not row[header_map['xoa_sua']]:
                errors.append(f'Dòng {idx}: Cột "xoa_sua" không được rỗng.')
                continue
            # Kiểm tra số lượng phải là số
            sl_ton = row[header_map['sl_ton_dau_ky']]
            try:
                if sl_ton is not None and sl_ton != '':
                    float(sl_ton)
            except Exception:
                errors.append(f'Dòng {idx}: Cột "sl_ton_dau_ky" phải là số.')
                continue
            # Kiểm tra đơn giá phải là số
            don_gia = row[header_map['don_gia_ton_dau_ky']]
            try:
                if don_gia is not None and don_gia != '':
                    float(don_gia)
            except Exception:
                errors.append(f'Dòng {idx}: Cột "don_gia_ton_dau_ky" phải là số.')
                continue
            # Nếu không có lỗi, tạo bản ghi
            LA_INVENTORY_CATEGORIES.objects.using(DATABASE_NAME).create(
                id_nhan_vien=row[header_map['id_nhan_vien']],
                xoa_sua=row[header_map['xoa_sua']],
                ma_hang=row[header_map['ma_hang']],
                ten_hang=row[header_map['ten_hang']],
                dvt=row[header_map['dvt']],
                sl_ton_dau_ky=sl_ton,
                don_gia_ton_dau_ky=don_gia,
                ma_kho_luu_tru=row[header_map['ma_kho_luu_tru']],
            )
            count += 1
        if errors:
            return Response({
                'message': f'Đã import {count} dòng thành công, nhưng có lỗi:',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': f'Import thành công {count} dòng!'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==============================================================================
# CLIENT CATEGORIES
# ==============================================================================

# Import bulk data
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bulk_data_LA_CLIENT_CATEGORIES(request):
    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if 'import-data' not in wb.sheetnames:
            return Response({'error': 'Không tìm thấy sheet "import-data" trong file Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['import-data']
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Sheet "import-data" không có dữ liệu.'}, status=status.HTTP_400_BAD_REQUEST)
        header = rows[0]
        header_map = {col: idx for idx, col in enumerate(header)}
        required_fields = ['id_nhan_vien', 'xoa_sua', 'ma_khach_hang', 'ten_khach_hang', 'ma_phan_loai_01', 'ma_phan_loai_02', 'ma_phan_loai_03', 'ma_phan_loai_04', 'ma_phan_loai_05', 'ma_phan_loai_06', 'ma_phan_loai_07', 'ma_phan_loai_08', 'mst', 'dia_chi']
        for field in required_fields:
            if field not in header_map:
                return Response({'error': f'Thiếu cột bắt buộc: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra mã khách hàng trùng trước khi import
        for idx, row in enumerate(rows[1:], start=2):
            ma_khach_hang = row[header_map['ma_khach_hang']]
            if ma_khach_hang and LA_CLIENT_CATEGORIES.objects.using(DATABASE_NAME).filter(ma_khach_hang=ma_khach_hang, xoa_sua="new").exists():
                return Response({'error': f'Mã khách hàng "{ma_khach_hang}" đã tồn tại.'}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            if not row or not row[header_map['ma_khach_hang']]:
                errors.append(f'Dòng {idx}: Cột "ma_khach_hang" không được rỗng.')
                continue
            if not row[header_map['xoa_sua']]:
                errors.append(f'Dòng {idx}: Cột "xoa_sua" không được rỗng.')
                continue
            LA_CLIENT_CATEGORIES.objects.using(DATABASE_NAME).create(
                id_nhan_vien=row[header_map['id_nhan_vien']],
                xoa_sua=row[header_map['xoa_sua']],
                ma_khach_hang=row[header_map['ma_khach_hang']],
                ten_khach_hang=row[header_map['ten_khach_hang']],
                ma_phan_loai_01=row[header_map['ma_phan_loai_01']],
                ma_phan_loai_02=row[header_map['ma_phan_loai_02']],
                ma_phan_loai_03=row[header_map['ma_phan_loai_03']],
                ma_phan_loai_04=row[header_map['ma_phan_loai_04']],
                ma_phan_loai_05=row[header_map['ma_phan_loai_05']],
                ma_phan_loai_06=row[header_map['ma_phan_loai_06']],
                ma_phan_loai_07=row[header_map['ma_phan_loai_07']],
                ma_phan_loai_08=row[header_map['ma_phan_loai_08']],
                mst=row[header_map['mst']],
                dia_chi=row[header_map['dia_chi']],
            )
            count += 1
        if errors:
            return Response({
                'message': f'Đã import {count} dòng thành công, nhưng có lỗi:',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': f'Import thành công {count} dòng!'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# Lazy loading pagination class
class ClientPagination(PageNumberPagination):
    page_size = 25  # Default number of records per page
    page_size_query_param = 'limit'  # Allow the client to specify the page size
    max_page_size = 100  # Maximum number of records per page

# Get all data
class get_data_LA_CLIENT_CATEGORIES(viewsets.ModelViewSet):
    queryset = LA_CLIENT_CATEGORIES.objects.using(DATABASE_NAME).filter(xoa_sua="new").order_by("-ma_khach_hang")
    serializer_class = LAClientCategoriesSerializer
    pagination_class = ClientPagination  # Add pagination support

    @classmethod
    def as_view(cls, actions=None, **initkwargs):
        if actions is None:
            actions = {'get': 'list'}
        return super().as_view(actions, **initkwargs)


# Create LA_CLIENT_CATEGORIES
class LAClientCategoriesCreateView(APIView):
    def post(self, request):
        data = request.data
        ma_khach_hang = data.get("ma_khach_hang")
        action = data.pop("action", None)  # Remove 'action' from data

        # Check if a record with ma_khach_hang exists
        existing_record = LA_CLIENT_CATEGORIES.objects.filter(ma_khach_hang=ma_khach_hang).first()

        if action == "create":
            if existing_record:
                if LA_CLIENT_CATEGORIES.objects.filter(ma_khach_hang=ma_khach_hang, xoa_sua="new").exists():
                    # Return an error if the record already exists with xoa_sua = "new"
                    return Response({"error": "Record with ma_khach_hang already exists and xoa_sua is 'new'."}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    # Create a new record with xoa_sua = "new"
                    data["xoa_sua"] = "new"
                    LA_CLIENT_CATEGORIES.objects.create(**data)
                    return Response({"message": "New record created successfully."}, status=status.HTTP_201_CREATED)                    
            else:
                # Create a new record since ma_khach_hang does not exist
                LA_CLIENT_CATEGORIES.objects.create(**data)
                return Response({"message": "Record created successfully."}, status=status.HTTP_201_CREATED)
        elif action == "edit":
            if existing_record:
                if existing_record.xoa_sua == "new":
                    # Update existing record's xoa_sua to "old"
                    existing_record.xoa_sua = "old"
                    existing_record.save()
                    # Create a new record with xoa_sua = "new"
                    data["xoa_sua"] = "new"
                    LA_CLIENT_CATEGORIES.objects.create(**data)
                    return Response({"message": "Record updated and new record created successfully."}, status=status.HTTP_200_OK)
                else:
                    # Update the existing record directly
                    for key, value in data.items():
                        setattr(existing_record, key, value)
                    existing_record.save()
                    return Response({"message": "Record updated successfully."}, status=status.HTTP_200_OK)
            else:
                # Return an error if the record does not exist
                return Response({"error": "Record with ma_khach_hang does not exist."}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"error": "Invalid action specified."}, status=status.HTTP_400_BAD_REQUEST)

# Get next ma_khach_hang
class GetNextMaKhachHangView(APIView):
    def get(self, request):
        # Get the latest ma_khach_hang in the format KH00000
        latest_record = LA_CLIENT_CATEGORIES.objects.filter(ma_khach_hang__startswith="KH").order_by("-ma_khach_hang").first()
        
        if latest_record:
            latest_ma_khach_hang = latest_record.ma_khach_hang
            # Extract the numeric part and increment it
            match = re.match(r"KH(\d+)", latest_ma_khach_hang)
            if match:
                next_number = int(match.group(1)) + 1
                next_ma_khach_hang = f"KH{next_number:05d}"
                return Response({"next_ma_khach_hang": next_ma_khach_hang}, status=status.HTTP_200_OK)
        
        # If no records exist, return KH00001
        return Response({"next_ma_khach_hang": "KH00001"}, status=status.HTTP_200_OK)

# Export LA_CLIENT_CATEGORIES to Excel
class ExportLAClientCategoriesToExcel(APIView):
    def get(self, request):
        # Create a workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "LA_CLIENT_CATEGORIES"

        # Define the headers
        headers = [
            "ID", "Date", "ID Nhân Viên", "Xóa/Sửa", "Mã Khách Hàng", "Tên Khách Hàng",
            "Mã Phân Loại 01", "Mã Phân Loại 02", "Mã Phân Loại 03", "Mã Phân Loại 04",
            "Mã Phân Loại 05", "Mã Phân Loại 06", "Mã Phân Loại 07", "Mã Phân Loại 08",
            "MST", "Địa Chỉ"
        ]
        worksheet.append(headers)

        # Fetch all data from the model
        clients = LA_CLIENT_CATEGORIES.objects.all()

        # Add data rows
        for client in clients:
            worksheet.append([
                str(client.id),  # Convert UUID to string
                client.date.replace(tzinfo=None),  # Remove timezone from datetime
                client.id_nhan_vien, client.xoa_sua, client.ma_khach_hang,
                client.ten_khach_hang, client.ma_phan_loai_01, client.ma_phan_loai_02,
                client.ma_phan_loai_03, client.ma_phan_loai_04, client.ma_phan_loai_05,
                client.ma_phan_loai_06, client.ma_phan_loai_07, client.ma_phan_loai_08,
                client.mst, client.dia_chi
            ])

        # Adjust column widths
        for col_num, col_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 20

        # Create a response with the Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="LA_CLIENT_CATEGORIES.xlsx"'
        workbook.save(response)
        return response

# Update xoa_sua field
class UpdateXoaSuaView(APIView):
    def post(self, request):
        ma_khach_hang = request.data.get("ma_khach_hang")
        pass_field = request.data.get("pass_field")
        if not pass_field or pass_field != "admincome":
            return Response({"error": "Invalid or missing password."}, status=status.HTTP_403_FORBIDDEN)

        if not ma_khach_hang:
            return Response({"error": "ma_khach_hang is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if the record exists with xoa_sua = "new"
        record = LA_CLIENT_CATEGORIES.objects.filter(ma_khach_hang=ma_khach_hang, xoa_sua="new").first()
        if not record:
            return Response({"error": "Record with ma_khach_hang does not exist or xoa_sua is not 'new'."}, status=status.HTTP_404_NOT_FOUND)

        # Check the time difference
        time_difference = now() - record.date
        if time_difference < timedelta(hours=0, minutes=2):
            # Update xoa_sua to "delete"
            record.xoa_sua = "delete"
            record.save()
            return Response({"message": "Record updated successfully."}, status=status.HTTP_200_OK)
        else:
            return Response({"error": "overtime to delete"}, status=status.HTTP_400_BAD_REQUEST)
