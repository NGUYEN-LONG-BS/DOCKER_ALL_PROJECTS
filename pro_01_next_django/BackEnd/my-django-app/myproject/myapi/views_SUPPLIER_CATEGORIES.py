from django.utils.timezone import now
from django.db import models

from rest_framework import status
from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.pagination import PageNumberPagination

from .models_TB import TB_SUPPLIER_CATEGORIES
from .models_LA import LA_SUPPLIER_CATEGORIES
from .models_PA import PA_SUPPLIER_CATEGORIES
from .models_Ha_Noi import HANOI_SUPPLIER_CATEGORIES
from .models_Mien_Tay import MIENTAY_SUPPLIER_CATEGORIES
from .models_Nam_An import NAMAN_SUPPLIER_CATEGORIES

from .serializers_LA import LASupplierCategoriesSerializer

import openpyxl
import re
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import timedelta

DATABASE_NAME_default = 'default'
DATABASE_NAME_tb = 'tb'

# Define model mappings
MODEL_MAP_SUPPLIER_CATEGORIES = {
    "TB": (TB_SUPPLIER_CATEGORIES, "tb"),
    "LA": (LA_SUPPLIER_CATEGORIES, "tala"),
    "PA": (PA_SUPPLIER_CATEGORIES, "pa"),
    "HANOI": (HANOI_SUPPLIER_CATEGORIES, "hanoi"),
    "MIENTAY": (MIENTAY_SUPPLIER_CATEGORIES, "mientay"),
    "NAMAN": (NAMAN_SUPPLIER_CATEGORIES, "naman"),
}

# ==========================================================================
# Import bulk data
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bulk_data_to_all_SUPPLIER_CATEGORIES(request):
    file_obj = request.FILES.get('file')
    model_key = request.data.get("model_key", "TB")
    model_tuple = MODEL_MAP_SUPPLIER_CATEGORIES.get(model_key)
    if not model_tuple:
        return Response({'error': 'Invalid model_key.'}, status=status.HTTP_400_BAD_REQUEST)
    ModelClass, db_name = model_tuple
    if not file_obj:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if 'import-data' not in wb.sheetnames:
            return Response({'error': 'Không tìm thấy sheet "import-data" trong file Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['import-data']
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Sheet "import-data" không có dữ liệu.'}, status=status.HTTP_400_BAD_REQUEST)
        header = rows[0]
        header_map = {col: idx for idx, col in enumerate(header)}
        required_fields = ['id_nhan_vien', 'xoa_sua', 'ma_nha_cung_cap', 'ten_nha_cung_cap', 'ma_phan_loai_01', 'ma_phan_loai_02', 'ma_phan_loai_03', 'ma_phan_loai_04', 'ma_phan_loai_05', 'ma_phan_loai_06', 'ma_phan_loai_07', 'ma_phan_loai_08', 'mst', 'dia_chi']
        for field in required_fields:
            if field not in header_map:
                return Response({'error': f'Thiếu cột bắt buộc: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra mã khách hàng trùng trước khi import
        for idx, row in enumerate(rows[1:], start=2):
            ma_nha_cung_cap = row[header_map['ma_nha_cung_cap']]
            if ma_nha_cung_cap and ModelClass.objects.using(db_name).filter(ma_nha_cung_cap=ma_nha_cung_cap, xoa_sua="new").exists():
                return Response({'error': f'Mã khách hàng "{ma_nha_cung_cap}" đã tồn tại.'}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            if not row or not row[header_map['ma_nha_cung_cap']]:
                errors.append(f'Dòng {idx}: Cột "ma_nha_cung_cap" không được rỗng.')
                continue
            if not row[header_map['xoa_sua']]:
                errors.append(f'Dòng {idx}: Cột "xoa_sua" không được rỗng.')
                continue
            ModelClass.objects.using(db_name).create(
                id_nhan_vien=row[header_map['id_nhan_vien']],
                xoa_sua=row[header_map['xoa_sua']],
                ma_nha_cung_cap=row[header_map['ma_nha_cung_cap']],
                ten_nha_cung_cap=row[header_map['ten_nha_cung_cap']],
                ma_phan_loai_01=row[header_map['ma_phan_loai_01']],
                ma_phan_loai_02=row[header_map['ma_phan_loai_02']],
                ma_phan_loai_03=row[header_map['ma_phan_loai_03']],
                ma_phan_loai_04=row[header_map['ma_phan_loai_04']],
                ma_phan_loai_05=row[header_map['ma_phan_loai_05']],
                ma_phan_loai_06=row[header_map['ma_phan_loai_06']],
                ma_phan_loai_07=row[header_map['ma_phan_loai_07']],
                ma_phan_loai_08=row[header_map['ma_phan_loai_08']],
                mst=row[header_map['mst']],
                dia_chi=row[header_map['dia_chi']],
            )
            count += 1
        if errors:
            return Response({
                'message': f'Đã import {count} dòng thành công, nhưng có lỗi:',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': f'Import thành công {count} dòng!'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==========================================================================
# Lazy loading pagination class
class SupplierPagination(PageNumberPagination):
    page_size = 25  # Default number of records per page
    page_size_query_param = 'limit'  # Allow the supplier to specify the page size
    max_page_size = 100  # Maximum number of records per page

# ==========================================================================
# Get all data
class get_data_TB_SUPPLIER_CATEGORIES(viewsets.ModelViewSet):
    queryset = TB_SUPPLIER_CATEGORIES.objects.using(DATABASE_NAME_tb).filter(xoa_sua="new").order_by("-ma_nha_cung_cap")
    serializer_class = LASupplierCategoriesSerializer
    pagination_class = SupplierPagination  # Add pagination support

    @classmethod
    def as_view(cls, actions=None, **initkwargs):
        if actions is None:
            actions = {'get': 'list'}
        return super().as_view(actions, **initkwargs)

# ==========================================================================
# Create TB_SUPPLIER_CATEGORIES
class TBSupplierCategoriesCreateView(APIView):
    def post(self, request):
        data = request.data
        ma_nha_cung_cap = data.get("ma_nha_cung_cap")
        action = data.pop("action", None)  # Remove 'action' from data

        # Check if a record with ma_nha_cung_cap exists
        existing_record = TB_SUPPLIER_CATEGORIES.objects.filter(ma_nha_cung_cap=ma_nha_cung_cap).first()

        if action == "create":
            if existing_record:
                if TB_SUPPLIER_CATEGORIES.objects.filter(ma_nha_cung_cap=ma_nha_cung_cap, xoa_sua="new").exists():
                    # Return an error if the record already exists with xoa_sua = "new"
                    return Response({"error": "Record with ma_nha_cung_cap already exists and xoa_sua is 'new'."}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    # Create a new record with xoa_sua = "new"
                    data["xoa_sua"] = "new"
                    TB_SUPPLIER_CATEGORIES.objects.create(**data)
                    return Response({"message": "New record created successfully."}, status=status.HTTP_201_CREATED)                    
            else:
                # Create a new record since ma_nha_cung_cap does not exist
                TB_SUPPLIER_CATEGORIES.objects.create(**data)
                return Response({"message": "Record created successfully."}, status=status.HTTP_201_CREATED)
        elif action == "edit":
            if existing_record:
                if existing_record.xoa_sua == "new":
                    # Update existing record's xoa_sua to "old"
                    existing_record.xoa_sua = "old"
                    existing_record.save()
                    # Create a new record with xoa_sua = "new"
                    data["xoa_sua"] = "new"
                    TB_SUPPLIER_CATEGORIES.objects.create(**data)
                    return Response({"message": "Record updated and new record created successfully."}, status=status.HTTP_200_OK)
                else:
                    # Update the existing record directly
                    for key, value in data.items():
                        setattr(existing_record, key, value)
                    existing_record.save()
                    return Response({"message": "Record updated successfully."}, status=status.HTTP_200_OK)
            else:
                # Return an error if the record does not exist
                return Response({"error": "Record with ma_nha_cung_cap does not exist."}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"error": "Invalid action specified."}, status=status.HTTP_400_BAD_REQUEST)

# ==========================================================================
# Get next ma_nha_cung_cap
@api_view(['GET'])
def get_next_ma_nha_cung_cap(request):
    model_key = request.GET.get('model_key', 'TB')
    model_tuple = MODEL_MAP_SUPPLIER_CATEGORIES.get(model_key)
    if not model_tuple:
        return Response({'error': 'Invalid model_key'}, status=status.HTTP_400_BAD_REQUEST)
    ModelClass, db_name = model_tuple
    latest_record = ModelClass.objects.using(db_name).filter(ma_nha_cung_cap__startswith="NCC").order_by("-ma_nha_cung_cap").first()
    if latest_record:
        latest_ma_nha_cung_cap = latest_record.ma_nha_cung_cap
        match = re.match(r"NCC(\d+)", latest_ma_nha_cung_cap)
        if match:
            next_number = int(match.group(1)) + 1
            next_ma_nha_cung_cap = f"NCC{next_number:05d}"
            return Response({"next_ma_nha_cung_cap": next_ma_nha_cung_cap}, status=status.HTTP_200_OK)
    return Response({"next_ma_nha_cung_cap": "NCC00001"}, status=status.HTTP_200_OK)

# ==========================================================================
# Export TB_SUPPLIER_CATEGORIES to Excel
class ExportTBSupplierCategoriesToExcel(APIView):
    def get(self, request):
        model_key = request.GET.get('model_key', 'TB')
        model_tuple = MODEL_MAP_SUPPLIER_CATEGORIES.get(model_key)
        if not model_tuple:
            return Response({'error': 'Invalid model_key'}, status=status.HTTP_400_BAD_REQUEST)
        ModelClass, db_name = model_tuple
        # Create a workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"{model_key}_SUPPLIER_CATEGORIES"

        # Define the headers
        headers = [
            "ID", "Date", "ID Nhân Viên", "Xóa/Sửa", "Mã Khách Hàng", "Tên Khách Hàng",
            "Mã Phân Loại 01", "Mã Phân Loại 02", "Mã Phân Loại 03", "Mã Phân Loại 04",
            "Mã Phân Loại 05", "Mã Phân Loại 06", "Mã Phân Loại 07", "Mã Phân Loại 08",
            "MST", "Địa Chỉ"
        ]
        worksheet.append(headers)

        # Fetch all data from the model
        suppliers = ModelClass.objects.using(db_name).all()

        # Add data rows
        for supplier in suppliers:
            worksheet.append([
                str(supplier.id),  # Convert UUID to string
                supplier.date.replace(tzinfo=None),  # Remove timezone from datetime
                supplier.id_nhan_vien, supplier.xoa_sua, supplier.ma_nha_cung_cap,
                supplier.ten_nha_cung_cap, supplier.ma_phan_loai_01, supplier.ma_phan_loai_02,
                supplier.ma_phan_loai_03, supplier.ma_phan_loai_04, supplier.ma_phan_loai_05,
                supplier.ma_phan_loai_06, supplier.ma_phan_loai_07, supplier.ma_phan_loai_08,
                supplier.mst, supplier.dia_chi
            ])

        # Adjust column widths
        for col_num, col_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 20

        # Create a response with the Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{model_key}_SUPPLIER_CATEGORIES.xlsx"'
        workbook.save(response)
        return response

# ==========================================================================
# Update xoa_sua field
class UpdateXoaSuaSupplierView(APIView):
    def post(self, request):
        ma_nha_cung_cap = request.data.get("ma_nha_cung_cap")
        pass_field = request.data.get("pass_field")
        if not pass_field or pass_field != "admincome":
            return Response({"error": "Invalid or missing password."}, status=status.HTTP_403_FORBIDDEN)

        if not ma_nha_cung_cap:
            return Response({"error": "ma_nha_cung_cap is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if the record exists with xoa_sua = "new"
        record = TB_SUPPLIER_CATEGORIES.objects.filter(ma_nha_cung_cap=ma_nha_cung_cap, xoa_sua="new").first()
        if not record:
            return Response({"error": "Record with ma_nha_cung_cap does not exist or xoa_sua is not 'new'."}, status=status.HTTP_404_NOT_FOUND)

        # Check the time difference
        time_difference = now() - record.date
        if time_difference < timedelta(hours=0, minutes=2):
            # Update xoa_sua to "delete"
            record.xoa_sua = "delete"
            record.save()
            return Response({"message": "Record updated successfully."}, status=status.HTTP_200_OK)
        else:
            return Response({"error": "overtime to delete"}, status=status.HTTP_400_BAD_REQUEST)

# ==========================================================================
# api to search supplier categories
@api_view(['GET'])
def search_supplier_categories(request):
    query = request.GET.get('q', '').strip()
    model_key = request.GET.get('model_key', 'TB')
    model_tuple = MODEL_MAP_SUPPLIER_CATEGORIES.get(model_key)
    if not model_tuple:
        return Response({'results': [], 'message': 'Invalid model_key'}, status=400)
    ModelClass, db_name = model_tuple
    if not query:
        return Response({'results': []})
    qs = ModelClass.objects.using(db_name).filter(
        models.Q(ma_nha_cung_cap__icontains=query) |
        models.Q(ten_nha_cung_cap__icontains=query) |
        models.Q(mst__icontains=query)
    )
    results = [
        {
            'ma_nha_cung_cap': c.ma_nha_cung_cap,
            'ten_nha_cung_cap': c.ten_nha_cung_cap,
            'mst': c.mst,
            'dia_chi': c.dia_chi
        }
        for c in qs[:20]
    ]
    if not results:
        return Response({'results': [], 'message': 'no data'})
    return Response({'results': results})

