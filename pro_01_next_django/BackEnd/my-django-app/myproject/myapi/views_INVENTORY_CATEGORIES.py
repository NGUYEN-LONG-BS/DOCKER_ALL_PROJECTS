from django.db import models
from django.utils import timezone

from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, parser_classes
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.generics import ListAPIView

from .models_TB import TB_INVENTORY_CATEGORIES
from .models_TB import TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED
from .models_LA import LA_INVENTORY_CATEGORIES
from .models_Ha_Noi import HANOI_INVENTORY_CATEGORIES
from .models_Mien_Tay import MIENTAY_INVENTORY_CATEGORIES
from .models_PA import PA_INVENTORY_CATEGORIES
from .models_Nam_An import NAMAN_INVENTORY_CATEGORIES

from .serializers_TB import TBInventoryCategoriesSerializer
from .serializers_LA import LAInventoryCategoriesSerializer
from .serializers_PA import PAInventoryCategoriesSerializer
from .serializers_Nam_An import NAMANInventoryCategoriesSerializer
from .serializers_Mien_Tay import MIENTAYInventoryCategoriesSerializer
from .serializers_Ha_Noi import HANOIInventoryCategoriesSerializer
import openpyxl
from datetime import datetime

MODEL_MAP_INVENTORY_CATEGORIES = {
    "null": ("null", "null", "null"),
    "TB": (TB_INVENTORY_CATEGORIES, TBInventoryCategoriesSerializer, "tb"),
    "LA": (LA_INVENTORY_CATEGORIES, LAInventoryCategoriesSerializer, "tala"),
    "PA": (PA_INVENTORY_CATEGORIES, PAInventoryCategoriesSerializer, "pa"),
    "NAMAN": (NAMAN_INVENTORY_CATEGORIES, NAMANInventoryCategoriesSerializer, "naman"),
    "HANOI": (HANOI_INVENTORY_CATEGORIES, HANOIInventoryCategoriesSerializer, "hanoi"),
    "MIENTAY": (MIENTAY_INVENTORY_CATEGORIES, MIENTAYInventoryCategoriesSerializer, "mientay"),
}

# ==========================================================================
# get inventory categories
class TBInventoryCategoriesView(ListAPIView):
    def get_queryset(self):
        model_key = self.request.query_params.get('model_key', 'TB')
        model_tuple = MODEL_MAP_INVENTORY_CATEGORIES.get(model_key)
        if not model_tuple:
            first_model = list(MODEL_MAP_INVENTORY_CATEGORIES.values())[0][0]
            return first_model.objects.none()
        ModelClass, _, db_name = model_tuple
        return ModelClass.objects.using(db_name).all()

    def get_serializer_class(self):
        model_key = self.request.query_params.get('model_key', 'TB')
        model_tuple = MODEL_MAP_INVENTORY_CATEGORIES.get(model_key)
        if not model_tuple:
            first_serializer = list(MODEL_MAP_INVENTORY_CATEGORIES.values())[0][1]
            return first_serializer
        _, SerializerClass, _ = model_tuple
        return SerializerClass

# ==========================================================================
# Import bulk data
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bulk_data_to_all_INVENTORY_CATEGORIES(request):
    file_obj = request.FILES.get('file')
    model_key = request.data.get("model_key", "TB")
    model_tuple = MODEL_MAP_INVENTORY_CATEGORIES.get(model_key)
    if not model_tuple:
        return Response({'error': 'Invalid model_key.'}, status=status.HTTP_400_BAD_REQUEST)
    ModelClass, db_name = model_tuple
    if not file_obj:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if 'import-data' not in wb.sheetnames:
            return Response({'error': 'Không tìm thấy sheet "import-data" trong file Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['import-data']
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Sheet "import-data" không có dữ liệu.'}, status=status.HTTP_400_BAD_REQUEST)
        header = rows[0]
        header_map = {col: idx for idx, col in enumerate(header)}
        required_fields = ['id_nhan_vien', 'xoa_sua', 'ma_hang', 'ten_hang', 'dvt', 'sl_ton_dau_ky', 'don_gia_ton_dau_ky', 'ma_kho_luu_tru']
        for field in required_fields:
            if field not in header_map:
                return Response({'error': f'Thiếu cột bắt buộc: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra mã hàng trùng trước khi import
        for idx, row in enumerate(rows[1:], start=2):
            ma_hang = row[header_map['ma_hang']]
            if ma_hang and ModelClass.objects.using(db_name).filter(ma_hang=ma_hang, xoa_sua="new").exists():
                return Response({'error': f'Mã hàng "{ma_hang}" đã tồn tại.'}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            # Kiểm tra dòng trống hoặc thiếu mã hàng
            if not row or not row[header_map['ma_hang']]:
                errors.append(f'Dòng {idx}: Cột "ma_hang" không được rỗng.')
                continue
            # Kiểm tra cột xoa_sua không được rỗng
            if not row[header_map['xoa_sua']]:
                errors.append(f'Dòng {idx}: Cột "xoa_sua" không được rỗng.')
                continue
            # Kiểm tra số lượng phải là số
            sl_ton = row[header_map['sl_ton_dau_ky']]
            try:
                if sl_ton is not None and sl_ton != '':
                    float(sl_ton)
            except Exception:
                errors.append(f'Dòng {idx}: Cột "sl_ton_dau_ky" phải là số.')
                continue
            # Kiểm tra đơn giá phải là số
            don_gia = row[header_map['don_gia_ton_dau_ky']]
            try:
                if don_gia is not None and don_gia != '':
                    float(don_gia)
            except Exception:
                errors.append(f'Dòng {idx}: Cột "don_gia_ton_dau_ky" phải là số.')
                continue
            ModelClass.objects.using(db_name).create(
                id_nhan_vien=row[header_map['id_nhan_vien']],
                xoa_sua=row[header_map['xoa_sua']],
                ma_hang=row[header_map['ma_hang']],
                ten_hang=row[header_map['ten_hang']],
                dvt=row[header_map['dvt']],
                sl_ton_dau_ky=sl_ton,
                don_gia_ton_dau_ky=don_gia,
                ma_kho_luu_tru=row[header_map['ma_kho_luu_tru']],
            )
            count += 1
        if errors:
            return Response({
                'message': f'Đã import {count} dòng thành công, nhưng có lỗi:',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': f'Import thành công {count} dòng!'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# ==========================================================================
# api to search inventory categories
@api_view(['GET'])
def search_inventory_categories(request):
    query = request.GET.get('q', '').strip()
    model_key = request.GET.get('model_key', 'TB')
    model_tuple = MODEL_MAP_INVENTORY_CATEGORIES.get(model_key)
    if not model_tuple:
        return Response({'results': [], 'message': 'Invalid model_key'}, status=400)
    ModelClass, _, db_name = model_tuple
    if not query:
        return Response({'results': []})
    qs = ModelClass.objects.using(db_name).filter(
        models.Q(ma_hang__icontains=query) |
        models.Q(ten_hang__icontains=query)
    )
    results = [
        {
            'ma_hang': c.ma_hang,
            'ten_hang': c.ten_hang,
            'dvt': getattr(c, 'dvt', None)
        }
        for c in qs[:20]
    ]
    if not results:
        return Response({'results': [], 'message': 'no data'})
    return Response({'results': results})

# ==========================================================================
# api to submit inventory categories
@api_view(['POST'])
def submit_inventory_categories(request):
    if request.method == 'POST':
        ma_hang = request.data.get('ma_hang')
        model_key = request.data.get('model_key', 'TB')
        model_tuple = MODEL_MAP_INVENTORY_CATEGORIES.get(model_key)
        if not model_tuple:
            return Response({'error': 'Invalid model_key.'}, status=status.HTTP_400_BAD_REQUEST)
        ModelClass, SerializerClass, db_name = model_tuple
        if not ma_hang:
            return Response({"error": "Trường 'ma_hang' không được để trống."}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra điều kiện: ma_hang chưa tồn tại hoặc đã tồn tại nhưng xoa_sua khác 'new'
        exists_new = ModelClass.objects.using(db_name).filter(ma_hang=ma_hang, xoa_sua="new").exists()
        if exists_new:
            return Response({"error": f"Mã hàng '{ma_hang}' đã tồn tại, không thể thêm mới."}, status=status.HTTP_400_BAD_REQUEST)
        serializer = SerializerClass(data=request.data)
        if serializer.is_valid():
            instance = ModelClass(**serializer.validated_data)
            instance.save(using=db_name)
            return Response({"message": "Dữ liệu đã được thêm thành công!"}, status=status.HTTP_201_CREATED)
        return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

# ==========================================================================
# api to check if ma_hang exists
class CheckMaHangExistView(APIView):
    def get(self, request, format=None):
        model_key = request.query_params.get('model_key', 'TB')
        model_tuple = MODEL_MAP_INVENTORY_CATEGORIES.get(model_key)
        if not model_tuple:
            return Response({'error': 'Invalid model_key'}, status=status.HTTP_400_BAD_REQUEST)
        ModelClass, _, db_name = model_tuple
        # Lấy giá trị cần tìm từ query parameters
        value_to_search = request.query_params.get('ma_hang', None)

        if not value_to_search:
            return Response(
                {'error': 'Mã hàng không được cung cấp'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Kiểm tra xem giá trị có tồn tại trong cột ma_hang của bảng không
        exists = ModelClass.objects.using(db_name).filter(ma_hang=value_to_search).exists()

        # Trả về kết quả với status 200
        return Response({'existed': exists}, status=status.HTTP_200_OK)


# ==========================================================================
# api to inventory report quantity

class InventoryReportQuantityView(APIView):
    def get(self, request, format=None):
        # Lấy ngày bắt đầu và kết thúc từ query param
        from_date_str = request.query_params.get('from_date')
        to_date_str = request.query_params.get('to_date')
        try:
            if from_date_str:
                from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
            else:
                from_date = timezone.now().date()
            if to_date_str:
                to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
            else:
                to_date = timezone.now().date()
        except Exception:
            return Response({'error': 'Sai định dạng ngày, dùng yyyy-mm-dd'}, status=400)

        # Lấy danh sách hàng hóa từ DB tb
        items = TB_INVENTORY_CATEGORIES.objects.using('tb').all()
        result = []

        for item in items:
            ma_hang = item.ma_hang

            # Tồn đầu kỳ: tổng nhập - tổng xuất trước from_date
            nhap_truoc = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.using('tb').filter(
                ma_hang=ma_hang,
                phan_loai_nhap_xuat_hoan='receipt',
                ngay_tren_phieu__date__lt=from_date
            ).aggregate(total=models.Sum('so_luong'))['total'] or 0

            xuat_truoc = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.using('tb').filter(
                ma_hang=ma_hang,
                phan_loai_nhap_xuat_hoan='issue',
                ngay_tren_phieu__date__lt=from_date
            ).aggregate(total=models.Sum('so_luong'))['total'] or 0

            ton_dau_ky = nhap_truoc - xuat_truoc

            # Nhập trong khoảng thời gian
            nhap_trong_khoang = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.using('tb').filter(
                ma_hang=ma_hang,
                phan_loai_nhap_xuat_hoan='receipt',
                ngay_tren_phieu__date__gte=from_date,
                ngay_tren_phieu__date__lte=to_date
            ).aggregate(total=models.Sum('so_luong'))['total'] or 0

            # Xuất trong khoảng thời gian
            xuat_trong_khoang = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.using('tb').filter(
                ma_hang=ma_hang,
                phan_loai_nhap_xuat_hoan='issue',
                ngay_tren_phieu__date__gte=from_date,
                ngay_tren_phieu__date__lte=to_date
            ).aggregate(total=models.Sum('so_luong'))['total'] or 0

            ton_cuoi_ky = ton_dau_ky + nhap_trong_khoang - xuat_trong_khoang

            result.append({
                'ma_hang': ma_hang,
                'ten_hang': item.ten_hang,
                'dvt': item.dvt,
                'ton_dau_ky': float(ton_dau_ky),
                'tong_nhap': float(nhap_trong_khoang),
                'tong_xuat': float(xuat_trong_khoang),
                'ton_cuoi_ky': float(ton_cuoi_ky),
            })

        return Response({
            'from_date': str(from_date),
            'to_date': str(to_date),
            'data': result
        }, status=200)