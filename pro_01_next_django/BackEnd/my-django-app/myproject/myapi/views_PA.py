from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
import openpyxl
from .models_PA import PA_INVENTORY_CATEGORIES

DATABASE_NAME = 'pa'

# ==============================================================================
# import inventory categories
# ==============================================================================
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bulk_data_PA_INVENTORY_CATEGORIES(request):
    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if 'import-data' not in wb.sheetnames:
            return Response({'error': 'Không tìm thấy sheet "import-data" trong file Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['import-data']
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Sheet "import-data" không có dữ liệu.'}, status=status.HTTP_400_BAD_REQUEST)
        header = rows[0]
        header_map = {col: idx for idx, col in enumerate(header)}
        required_fields = ['id_nhan_vien', 'xoa_sua', 'ma_hang', 'ten_hang', 'dvt', 'sl_ton_dau_ky', 'don_gia_ton_dau_ky', 'ma_kho_luu_tru']
        for field in required_fields:
            if field not in header_map:
                return Response({'error': f'Thiếu cột bắt buộc: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra mã hàng trùng trước khi import
        for idx, row in enumerate(rows[1:], start=2):
            ma_hang = row[header_map['ma_hang']]
            # xoa_sua = row[header_map['xoa_sua']]
            if ma_hang and PA_INVENTORY_CATEGORIES.objects.using(DATABASE_NAME).filter(ma_hang=ma_hang, xoa_sua="new").exists():
                return Response({'error': f'Mã hàng "{ma_hang}" ở dòng {idx} đã tồn tại, vui lòng kiểm tra lại.'}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            # Kiểm tra dòng trống hoặc thiếu mã hàng
            if not row or not row[header_map['ma_hang']]:
                errors.append(f'Dòng {idx}: Cột "ma_hang" không được rỗng.')
                continue
            # Kiểm tra cột xoa_sua không được rỗng
            if not row[header_map['xoa_sua']]:
                errors.append(f'Dòng {idx}: Cột "xoa_sua" không được rỗng.')
                continue
            # Kiểm tra số lượng phải là số
            sl_ton = row[header_map['sl_ton_dau_ky']]
            try:
                if sl_ton is not None and sl_ton != '':
                    float(sl_ton)
            except Exception:
                errors.append(f'Dòng {idx}: Cột "sl_ton_dau_ky" phải là số.')
                continue
            # Kiểm tra đơn giá phải là số
            don_gia = row[header_map['don_gia_ton_dau_ky']]
            try:
                if don_gia is not None and don_gia != '':
                    float(don_gia)
            except Exception:
                errors.append(f'Dòng {idx}: Cột "don_gia_ton_dau_ky" phải là số.')
                continue
            # Nếu không có lỗi, tạo bản ghi
            PA_INVENTORY_CATEGORIES.objects.using(DATABASE_NAME).create(
                id_nhan_vien=row[header_map['id_nhan_vien']],
                xoa_sua=row[header_map['xoa_sua']],
                ma_hang=row[header_map['ma_hang']],
                ten_hang=row[header_map['ten_hang']],
                dvt=row[header_map['dvt']],
                sl_ton_dau_ky=sl_ton,
                don_gia_ton_dau_ky=don_gia,
                ma_kho_luu_tru=row[header_map['ma_kho_luu_tru']],
            )
            count += 1
        if errors:
            return Response({
                'message': f'Đã import {count} dòng thành công, nhưng có lỗi:',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': f'Import thành công {count} dòng!'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)