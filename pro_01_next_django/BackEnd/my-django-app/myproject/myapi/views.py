from django.shortcuts import render
from django.http import JsonResponse
from django.http import FileResponse, HttpResponseNotFound
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now

from rest_framework import status
from rest_framework import viewsets
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.generics import ListAPIView
from rest_framework.pagination import PageNumberPagination

from .models import FormSubmission
from .models import LoginInfo, UserPermission
from .models import TB_INVENTORY_CATEGORIES, TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED
from .models import TB_CLIENT_CATEGORIES
from .models_LA import LA_CLIENT_CATEGORIES

from .serializers import FormSubmissionSerializer
from .serializers import LoginInfoSerializer, TBInventoryCategoriesSerializer, InventoryCategoriesSerializer
from .serializers import LoginInfoSerializer
from .serializers import TBInventoryCategoriesSerializer
from .serializers import UserPermissionSerializer, TB_CLIENT_CATEGORIES_Serializer
from .serializers import InventoryStockReceivedIssuedReturnedSerializer
from .serializers import InventoryStockSerializer
from .serializers import TBClientCategoriesSerializer

import json
import openpyxl
import os
import re
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import timedelta

DATABASE_NAME = 'default'

#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================

class FormSubmissionView(APIView):
    def post(self, request):
        try:
            serializer = FormSubmissionSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()  # Lưu dữ liệu vào database
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            # Ghi lại lỗi chi tiết để giúp bạn debug
            print(f"Error: {e}")
            return Response({"error": "Đã có lỗi xảy ra, vui lòng thử lại!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================

@api_view(['POST'])
def submit_login_info(request):
    if request.method == 'POST':
        serializer = LoginInfoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Dữ liệu đã được thêm thành công!"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def submit_inventory_categories(request):
    if request.method == 'POST':
        ma_hang = request.data.get('ma_hang')
        if not ma_hang:
            return Response({"error": "Trường 'ma_hang' không được để trống."}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra điều kiện: ma_hang chưa tồn tại hoặc đã tồn tại nhưng xoa_sua khác 'new'
        exists_new = TB_INVENTORY_CATEGORIES.objects.filter(ma_hang=ma_hang, xoa_sua="new").exists()
        if exists_new:
            return Response({"error": f"Mã hàng '{ma_hang}' đã tồn tại, không thể thêm mới."}, status=status.HTTP_400_BAD_REQUEST)
        serializer = InventoryCategoriesSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"message": "Dữ liệu đã được thêm thành công!"}, status=status.HTTP_201_CREATED)
        return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class LoginInfoListView(ListAPIView):
    queryset = LoginInfo.objects.all()
    serializer_class = LoginInfoSerializer

class TBInventoryCategoriesView(ListAPIView):
    queryset = TB_INVENTORY_CATEGORIES.objects.all()
    serializer_class = TBInventoryCategoriesSerializer

#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
#========================================================================================================================
    
def get_json_data(request):
    # Đường dẫn đến file JSON trong thư mục static/templates/json
    json_file_path = os.path.join(settings.BASE_DIR, 'static', 'templates', 'json', 'VT_QUAN_LY_HANG_HOA', 'inventoies_categoried.json')
    
    try:
        # Mở và đọc file JSON với mã hóa UTF-8
        with open(json_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        
        # Trả dữ liệu dưới dạng JSON
        return JsonResponse(data, safe=False)
    
    except FileNotFoundError:
        return JsonResponse({"error": "File not found"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON format"}, status=400)
    except UnicodeDecodeError:
        return JsonResponse({"error": "Unicode decoding error in file"}, status=400)

# ==============================================================================
# Save inventory
# ==============================================================================

class InventoryStockReceivedIssuedReturnedView(generics.ListCreateAPIView):
    queryset = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.all()
    serializer_class = InventoryStockReceivedIssuedReturnedSerializer
    
    # Override create method to handle multiple objects in a single request
    def create(self, request, *args, **kwargs):
        # Check if the request body contains a list of objects
        if isinstance(request.data, list):
            # If it's a list, we need to serialize each item
            serializer = self.get_serializer(data=request.data, many=True)
        else:
            # Otherwise, we process it as a single object
            serializer = self.get_serializer(data=request.data)

        # Validate the data
        if serializer.is_valid():
            # Save and return response
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# ==============================================================================
# Download file
# ==============================================================================
# myproject/views.py

def download_file_PRINT_TEMPLATE(request):
    # Đường dẫn tới file Excel
    file_path = os.path.join(settings.BASE_DIR, 'static/templates/download/PRINT_TEMPLATE.xlsx')
    
    # Kiểm tra nếu file tồn tại
    if os.path.exists(file_path):
        # Trả về file cho người dùng
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='PRINT_TEMPLATE.xlsx')
    else:
        # Nếu file không tồn tại, trả về lỗi 404
        return HttpResponseNotFound("File not found")

def download_file_IMPORT_TEMPLATE(request):
    # Đường dẫn tới file Excel
    file_path = os.path.join(settings.BASE_DIR, 'static/templates/download/IMPORT_TEMPLATE.xlsx')
    
    # Kiểm tra nếu file tồn tại
    if os.path.exists(file_path):
        # Trả về file cho người dùng
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename='IMPORT_TEMPLATE.xlsx')
    else:
        # Nếu file không tồn tại, trả về lỗi 404
        return HttpResponseNotFound("File not found")

# ==============================================================================
# Upload file
# ==============================================================================

@api_view(['POST'])
def import_data(request):
    print("Bắt đầu lưu file")

    # Lấy file từ request
    file = request.FILES.get('file')
    
    if not file:
        return Response({'error': 'No file uploaded'}, status=400)

    # Đảm bảo thư mục static/templates/upload tồn tại
    upload_dir = os.path.join(settings.BASE_DIR, 'static/templates/upload')
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)
        print(f"Thư mục {upload_dir} đã được tạo.")

    # Đặt tên file và kiểm tra tên file để đảm bảo không có ký tự đặc biệt
    file_name = file.name
    file_path = os.path.join(upload_dir, file_name)

    try:
        # Lưu file vào thư mục upload
        with open(file_path, 'wb') as f:
            for chunk in file.chunks():
                f.write(chunk)
        
        # In ra file path để kiểm tra
        print(f"File đã được lưu tại {file_path}")

        return Response({'message': 'File uploaded and saved successfully'}, status=200)
    
    except Exception as e:
        # Xử lý lỗi khi lưu file
        print(f"Lỗi khi lưu file: {str(e)}")
        return Response({'error': f'Error saving file: {str(e)}'}, status=500)

# ==============================================================================
# Get max number of slip
# ==============================================================================

class MaxSoPhieuView(APIView):

    def get(self, request, format=None):
        # Lấy tất cả các số phiếu từ bảng
        phieu_list = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.all()
        
        # Danh sách chứa các số phiếu và 6 số cuối của chúng
        max_phieu = None
        max_last_six = -1  # Giá trị ban đầu thấp nhất để so sánh

        for phieu in phieu_list:
            # Trích xuất 6 ký tự cuối và chuyển thành số
            last_six_digits = phieu.so_phieu[-6:]

            try:
                # So sánh các số cuối cùng (chuyển sang kiểu số nguyên để so sánh)
                last_six_number = int(last_six_digits)
                
                if last_six_number > max_last_six:
                    max_last_six = last_six_number
                    max_phieu = phieu.so_phieu
            except ValueError:
                continue  # Nếu không thể chuyển thành số thì bỏ qua

        if max_phieu:
            # Tăng số cuối lên 1
            max_last_six += 1
            
            # Tạo số phiếu mới
            new_number_slip = f"{max_phieu[:-6]}{max_last_six:06d}"
            
            # Trả về số phiếu moi
            return Response({'new_number_slip': new_number_slip}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Không có số phiếu nào'}, status=status.HTTP_404_NOT_FOUND)

# ==============================================================================
# Check number slip exist
# ==============================================================================
        
class CheckSoPhieuExistView(APIView):
    def get(self, request, format=None):
        # Lấy số phiếu từ query parameters
        value_to_search = request.query_params.get('so_phieu', None)

        if not value_to_search:
            return Response(
                {'error': 'Số phiếu không được cung cấp'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Kiểm tra xem số phiếu đã tồn tại trong cơ sở dữ liệu chưa
        exists = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.filter(so_phieu=value_to_search).exists()

        # Trả về kết quả với status 200, bất kể số phiếu tồn tại hay không
        return Response({'existed': exists}, status=status.HTTP_200_OK)
    
class CheckMaHangExistView(APIView):
    def get(self, request, format=None):
        # Lấy giá trị cần tìm từ query parameters
        value_to_search = request.query_params.get('ma_hang', None)

        if not value_to_search:
            return Response(
                {'error': 'Mã hàng không được cung cấp'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Kiểm tra xem giá trị có tồn tại trong cột ma_hang của bảng không
        exists = TB_INVENTORY_CATEGORIES.objects.filter(ma_hang=value_to_search).exists()

        # Trả về kết quả với status 200
        return Response({'existed': exists}, status=status.HTTP_200_OK)

# ==============================================================================
# List inventory stock
# ==============================================================================

class InventoryStockListView(APIView):
    def get(self, request, format=None):
        from_date = request.query_params.get('from_date')
        to_date = request.query_params.get('to_date')
        so_phieu = request.query_params.get('so_phieu')
        so_phieu_de_nghi = request.query_params.get('so_phieu_de_nghi')
        ma_doi_tuong = request.query_params.get('ma_doi_tuong')
        ma_hang = request.query_params.get('ma_hang')

        qs = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.all()

        # Đúng tên trường ngày
        if from_date:
            qs = qs.filter(ngay_tren_phieu__date__gte=from_date)
        if to_date:
            qs = qs.filter(ngay_tren_phieu__date__lte=to_date)
        if so_phieu:
            qs = qs.filter(so_phieu__icontains=so_phieu)
        if so_phieu_de_nghi:
            qs = qs.filter(so_phieu_de_nghi__icontains=so_phieu_de_nghi)
        if ma_doi_tuong:
            qs = qs.filter(ma_doi_tuong__icontains=ma_doi_tuong)
        if ma_hang:
            qs = qs.filter(ma_hang__icontains=ma_hang)

        qs = qs.order_by('-so_phieu')

        serialized_data = []
        for index, item in enumerate(qs):
            serializer = InventoryStockSerializer(item, context={'index': index})
            serialized_data.append(serializer.data)
        return Response(serialized_data, status=status.HTTP_200_OK)
    
# ==============================================================================
# check login
# ==============================================================================

@csrf_exempt
def check_login(request):
    if request.method == 'POST':
        try:
            # Lấy dữ liệu từ request body
            data = json.loads(request.body)
            login_id = data.get('login_id')
            pass_field = data.get('pass_field')

            # Kiểm tra xem login_id và pass_field có được cung cấp hay không
            if not login_id or not pass_field:
                return JsonResponse({'error': 'Vui lòng cung cấp login_id và pass_field'}, status=400)

            # Kiểm tra trong database
            exists = LoginInfo.objects.filter(login_id=login_id, pass_field=pass_field).exists()
            
            return JsonResponse({'result': exists}, status=200)
        
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Dữ liệu JSON không hợp lệ'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Phương thức không được hỗ trợ'}, status=405)

# ==============================================================================
# inherit slip
# ==============================================================================

class InventoryStockBySoPhieuView(APIView):
    def get(self, request, format=None):
        so_phieu = request.query_params.get('so_phieu')
        if not so_phieu:
            return Response({'error': 'Thiếu tham số so_phieu'}, status=status.HTTP_400_BAD_REQUEST)
        qs = TB_INVENTORY_STOCK_RECEIVED_ISSSUED_RETURNED.objects.filter(so_phieu=so_phieu, xoa_sua="new")
        serialized_data = []
        for index, item in enumerate(qs):
            serializer = InventoryStockSerializer(item, context={'index': index})
            serialized_data.append(serializer.data)
        return Response(serialized_data, status=status.HTTP_200_OK)

# ==============================================================================
# INVENTORY CATEGORIES
# ==============================================================================

# Import bulk data
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bulk_data_TB_INVENTORY_CATEGORIES(request):
    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if 'import-data' not in wb.sheetnames:
            return Response({'error': 'Không tìm thấy sheet "import-data" trong file Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['import-data']
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Sheet "import-data" không có dữ liệu.'}, status=status.HTTP_400_BAD_REQUEST)
        header = rows[0]
        header_map = {col: idx for idx, col in enumerate(header)}
        required_fields = ['id_nhan_vien', 'xoa_sua', 'ma_hang', 'ten_hang', 'dvt', 'sl_ton_dau_ky', 'don_gia_ton_dau_ky', 'ma_kho_luu_tru']
        for field in required_fields:
            if field not in header_map:
                return Response({'error': f'Thiếu cột bắt buộc: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra mã hàng trùng trước khi import
        for idx, row in enumerate(rows[1:], start=2):
            ma_hang = row[header_map['ma_hang']]
            # xoa_sua = row[header_map['xoa_sua']]
            if ma_hang and TB_INVENTORY_CATEGORIES.objects.using(DATABASE_NAME).filter(ma_hang=ma_hang, xoa_sua="new").exists():
                return Response({'error': f'Mã hàng "{ma_hang}" đã tồn tại.'}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            # Kiểm tra dòng trống hoặc thiếu mã hàng
            if not row or not row[header_map['ma_hang']]:
                errors.append(f'Dòng {idx}: Cột "ma_hang" không được rỗng.')
                continue
            # Kiểm tra cột xoa_sua không được rỗng
            if not row[header_map['xoa_sua']]:
                errors.append(f'Dòng {idx}: Cột "xoa_sua" không được rỗng.')
                continue
            # Kiểm tra số lượng phải là số
            sl_ton = row[header_map['sl_ton_dau_ky']]
            try:
                if sl_ton is not None and sl_ton != '':
                    float(sl_ton)
            except Exception:
                errors.append(f'Dòng {idx}: Cột "sl_ton_dau_ky" phải là số.')
                continue
            # Kiểm tra đơn giá phải là số
            don_gia = row[header_map['don_gia_ton_dau_ky']]
            try:
                if don_gia is not None and don_gia != '':
                    float(don_gia)
            except Exception:
                errors.append(f'Dòng {idx}: Cột "don_gia_ton_dau_ky" phải là số.')
                continue
            # Nếu không có lỗi, tạo bản ghi
            TB_INVENTORY_CATEGORIES.objects.using(DATABASE_NAME).create(
                id_nhan_vien=row[header_map['id_nhan_vien']],
                xoa_sua=row[header_map['xoa_sua']],
                ma_hang=row[header_map['ma_hang']],
                ten_hang=row[header_map['ten_hang']],
                dvt=row[header_map['dvt']],
                sl_ton_dau_ky=sl_ton,
                don_gia_ton_dau_ky=don_gia,
                ma_kho_luu_tru=row[header_map['ma_kho_luu_tru']],
            )
            count += 1
        if errors:
            return Response({
                'message': f'Đã import {count} dòng thành công, nhưng có lỗi:',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': f'Import thành công {count} dòng!'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)



# ==============================================================================
# Login and Permission
# ==============================================================================

class UserPermissionViewSet(viewsets.ModelViewSet):
    serializer_class = UserPermissionSerializer

    def get_queryset(self):
        queryset = UserPermission.objects.using(DATABASE_NAME).all().order_by('user_id')
        # Đánh lại số thứ tự (STT) cho từng bản ghi
        for idx, obj in enumerate(queryset, start=1):
            obj.stt = idx
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        # Thêm trường STT vào từng bản ghi trả về
        data = serializer.data
        for idx, item in enumerate(data, start=1):
            item['stt'] = idx
        return Response(data)
    
@api_view(['GET'])
def get_user_permission_info(request):
    user_id = request.query_params.get('user_id')
    if not user_id:
        return Response({'error': 'Thiếu user_id'}, status=400)
    user_permissions = UserPermission.objects.using(DATABASE_NAME).filter(user_id=user_id)
    if not user_permissions.exists():
        return Response({'error': 'Không tìm thấy user_id này'}, status=404)
    data = [
        {
            'subsidiary': up.subsidiary,
            'department': up.department
        }
        for up in user_permissions
    ]
    return Response(data, status=200)

# ==============================================================================
# TB_CLIENT_CATEGORIES
# ==============================================================================

# Import bulk data
MODEL_MAP = {
    "TB": (TB_CLIENT_CATEGORIES, "default"),
    "LA": (LA_CLIENT_CATEGORIES, "tala"),
}

@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bulk_data_to_all_CLIENT_CATEGORIES(request):
    file_obj = request.FILES.get('file')
    model_key = request.data.get("model_key", "TB")
    model_tuple = MODEL_MAP.get(model_key)
    if not model_tuple:
        return Response({'error': 'Invalid model_key.'}, status=status.HTTP_400_BAD_REQUEST)
    ModelClass, db_name = model_tuple
    if not file_obj:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if 'import-data' not in wb.sheetnames:
            return Response({'error': 'Không tìm thấy sheet "import-data" trong file Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['import-data']
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Sheet "import-data" không có dữ liệu.'}, status=status.HTTP_400_BAD_REQUEST)
        header = rows[0]
        header_map = {col: idx for idx, col in enumerate(header)}
        required_fields = ['id_nhan_vien', 'xoa_sua', 'ma_khach_hang', 'ten_khach_hang', 'ma_phan_loai_01', 'ma_phan_loai_02', 'ma_phan_loai_03', 'ma_phan_loai_04', 'ma_phan_loai_05', 'ma_phan_loai_06', 'ma_phan_loai_07', 'ma_phan_loai_08', 'mst', 'dia_chi']
        for field in required_fields:
            if field not in header_map:
                return Response({'error': f'Thiếu cột bắt buộc: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra mã khách hàng trùng trước khi import
        for idx, row in enumerate(rows[1:], start=2):
            ma_khach_hang = row[header_map['ma_khach_hang']]
            if ma_khach_hang and ModelClass.objects.using(db_name).filter(ma_khach_hang=ma_khach_hang, xoa_sua="new").exists():
                return Response({'error': f'Mã khách hàng "{ma_khach_hang}" đã tồn tại.'}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            if not row or not row[header_map['ma_khach_hang']]:
                errors.append(f'Dòng {idx}: Cột "ma_khach_hang" không được rỗng.')
                continue
            if not row[header_map['xoa_sua']]:
                errors.append(f'Dòng {idx}: Cột "xoa_sua" không được rỗng.')
                continue
            ModelClass.objects.using(db_name).create(
                id_nhan_vien=row[header_map['id_nhan_vien']],
                xoa_sua=row[header_map['xoa_sua']],
                ma_khach_hang=row[header_map['ma_khach_hang']],
                ten_khach_hang=row[header_map['ten_khach_hang']],
                ma_phan_loai_01=row[header_map['ma_phan_loai_01']],
                ma_phan_loai_02=row[header_map['ma_phan_loai_02']],
                ma_phan_loai_03=row[header_map['ma_phan_loai_03']],
                ma_phan_loai_04=row[header_map['ma_phan_loai_04']],
                ma_phan_loai_05=row[header_map['ma_phan_loai_05']],
                ma_phan_loai_06=row[header_map['ma_phan_loai_06']],
                ma_phan_loai_07=row[header_map['ma_phan_loai_07']],
                ma_phan_loai_08=row[header_map['ma_phan_loai_08']],
                mst=row[header_map['mst']],
                dia_chi=row[header_map['dia_chi']],
            )
            count += 1
        if errors:
            return Response({
                'message': f'Đã import {count} dòng thành công, nhưng có lỗi:',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': f'Import thành công {count} dòng!'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# Import bulk data
@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def import_bulk_data_TB_CLIENT_CATEGORIES(request):
    file_obj = request.FILES.get('file')
    if not file_obj:
        return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        if 'import-data' not in wb.sheetnames:
            return Response({'error': 'Không tìm thấy sheet "import-data" trong file Excel.'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['import-data']
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return Response({'error': 'Sheet "import-data" không có dữ liệu.'}, status=status.HTTP_400_BAD_REQUEST)
        header = rows[0]
        header_map = {col: idx for idx, col in enumerate(header)}
        required_fields = ['id_nhan_vien', 'xoa_sua', 'ma_khach_hang', 'ten_khach_hang', 'ma_phan_loai_01', 'ma_phan_loai_02', 'ma_phan_loai_03', 'ma_phan_loai_04', 'ma_phan_loai_05', 'ma_phan_loai_06', 'ma_phan_loai_07', 'ma_phan_loai_08', 'mst', 'dia_chi']
        for field in required_fields:
            if field not in header_map:
                return Response({'error': f'Thiếu cột bắt buộc: {field}'}, status=status.HTTP_400_BAD_REQUEST)
        # Kiểm tra mã khách hàng trùng trước khi import
        for idx, row in enumerate(rows[1:], start=2):
            ma_khach_hang = row[header_map['ma_khach_hang']]
            if ma_khach_hang and TB_CLIENT_CATEGORIES.objects.using(DATABASE_NAME).filter(ma_khach_hang=ma_khach_hang, xoa_sua="new").exists():
                return Response({'error': f'Mã khách hàng "{ma_khach_hang}" đã tồn tại.'}, status=status.HTTP_400_BAD_REQUEST)
        count = 0
        errors = []
        for idx, row in enumerate(rows[1:], start=2):
            if not row or not row[header_map['ma_khach_hang']]:
                errors.append(f'Dòng {idx}: Cột "ma_khach_hang" không được rỗng.')
                continue
            if not row[header_map['xoa_sua']]:
                errors.append(f'Dòng {idx}: Cột "xoa_sua" không được rỗng.')
                continue
            TB_CLIENT_CATEGORIES.objects.using(DATABASE_NAME).create(
                id_nhan_vien=row[header_map['id_nhan_vien']],
                xoa_sua=row[header_map['xoa_sua']],
                ma_khach_hang=row[header_map['ma_khach_hang']],
                ten_khach_hang=row[header_map['ten_khach_hang']],
                ma_phan_loai_01=row[header_map['ma_phan_loai_01']],
                ma_phan_loai_02=row[header_map['ma_phan_loai_02']],
                ma_phan_loai_03=row[header_map['ma_phan_loai_03']],
                ma_phan_loai_04=row[header_map['ma_phan_loai_04']],
                ma_phan_loai_05=row[header_map['ma_phan_loai_05']],
                ma_phan_loai_06=row[header_map['ma_phan_loai_06']],
                ma_phan_loai_07=row[header_map['ma_phan_loai_07']],
                ma_phan_loai_08=row[header_map['ma_phan_loai_08']],
                mst=row[header_map['mst']],
                dia_chi=row[header_map['dia_chi']],
            )
            count += 1
        if errors:
            return Response({
                'message': f'Đã import {count} dòng thành công, nhưng có lỗi:',
                'errors': errors
            }, status=status.HTTP_400_BAD_REQUEST)
        return Response({'message': f'Import thành công {count} dòng!'}, status=status.HTTP_201_CREATED)
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# Lazy loading pagination class
class ClientPagination(PageNumberPagination):
    page_size = 25  # Default number of records per page
    page_size_query_param = 'limit'  # Allow the client to specify the page size
    max_page_size = 100  # Maximum number of records per page

# Get all data
class get_data_TB_CLIENT_CATEGORIES(viewsets.ModelViewSet):
    queryset = TB_CLIENT_CATEGORIES.objects.using(DATABASE_NAME).filter(xoa_sua="new").order_by("-ma_khach_hang")
    serializer_class = TBClientCategoriesSerializer
    pagination_class = ClientPagination  # Add pagination support

    @classmethod
    def as_view(cls, actions=None, **initkwargs):
        if actions is None:
            actions = {'get': 'list'}
        return super().as_view(actions, **initkwargs)


# Create TB_CLIENT_CATEGORIES
class TBClientCategoriesCreateView(APIView):
    def post(self, request):
        data = request.data
        ma_khach_hang = data.get("ma_khach_hang")
        action = data.pop("action", None)  # Remove 'action' from data

        # Check if a record with ma_khach_hang exists
        existing_record = TB_CLIENT_CATEGORIES.objects.filter(ma_khach_hang=ma_khach_hang).first()

        if action == "create":
            if existing_record:
                if TB_CLIENT_CATEGORIES.objects.filter(ma_khach_hang=ma_khach_hang, xoa_sua="new").exists():
                    # Return an error if the record already exists with xoa_sua = "new"
                    return Response({"error": "Record with ma_khach_hang already exists and xoa_sua is 'new'."}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    # Create a new record with xoa_sua = "new"
                    data["xoa_sua"] = "new"
                    TB_CLIENT_CATEGORIES.objects.create(**data)
                    return Response({"message": "New record created successfully."}, status=status.HTTP_201_CREATED)                    
            else:
                # Create a new record since ma_khach_hang does not exist
                TB_CLIENT_CATEGORIES.objects.create(**data)
                return Response({"message": "Record created successfully."}, status=status.HTTP_201_CREATED)
        elif action == "edit":
            if existing_record:
                if existing_record.xoa_sua == "new":
                    # Update existing record's xoa_sua to "old"
                    existing_record.xoa_sua = "old"
                    existing_record.save()
                    # Create a new record with xoa_sua = "new"
                    data["xoa_sua"] = "new"
                    TB_CLIENT_CATEGORIES.objects.create(**data)
                    return Response({"message": "Record updated and new record created successfully."}, status=status.HTTP_200_OK)
                else:
                    # Update the existing record directly
                    for key, value in data.items():
                        setattr(existing_record, key, value)
                    existing_record.save()
                    return Response({"message": "Record updated successfully."}, status=status.HTTP_200_OK)
            else:
                # Return an error if the record does not exist
                return Response({"error": "Record with ma_khach_hang does not exist."}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({"error": "Invalid action specified."}, status=status.HTTP_400_BAD_REQUEST)

# Get next ma_khach_hang
class GetNextMaKhachHangView(APIView):
    def get(self, request):
        # Get the latest ma_khach_hang in the format KH00000
        latest_record = TB_CLIENT_CATEGORIES.objects.filter(ma_khach_hang__startswith="KH").order_by("-ma_khach_hang").first()
        
        if latest_record:
            latest_ma_khach_hang = latest_record.ma_khach_hang
            # Extract the numeric part and increment it
            match = re.match(r"KH(\d+)", latest_ma_khach_hang)
            if match:
                next_number = int(match.group(1)) + 1
                next_ma_khach_hang = f"KH{next_number:05d}"
                return Response({"next_ma_khach_hang": next_ma_khach_hang}, status=status.HTTP_200_OK)
        
        # If no records exist, return KH00001
        return Response({"next_ma_khach_hang": "KH00001"}, status=status.HTTP_200_OK)

# Export TB_CLIENT_CATEGORIES to Excel
class ExportTBClientCategoriesToExcel(APIView):
    def get(self, request):
        # Create a workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "TB_CLIENT_CATEGORIES"

        # Define the headers
        headers = [
            "ID", "Date", "ID Nhân Viên", "Xóa/Sửa", "Mã Khách Hàng", "Tên Khách Hàng",
            "Mã Phân Loại 01", "Mã Phân Loại 02", "Mã Phân Loại 03", "Mã Phân Loại 04",
            "Mã Phân Loại 05", "Mã Phân Loại 06", "Mã Phân Loại 07", "Mã Phân Loại 08",
            "MST", "Địa Chỉ"
        ]
        worksheet.append(headers)

        # Fetch all data from the model
        clients = TB_CLIENT_CATEGORIES.objects.all()

        # Add data rows
        for client in clients:
            worksheet.append([
                str(client.id),  # Convert UUID to string
                client.date.replace(tzinfo=None),  # Remove timezone from datetime
                client.id_nhan_vien, client.xoa_sua, client.ma_khach_hang,
                client.ten_khach_hang, client.ma_phan_loai_01, client.ma_phan_loai_02,
                client.ma_phan_loai_03, client.ma_phan_loai_04, client.ma_phan_loai_05,
                client.ma_phan_loai_06, client.ma_phan_loai_07, client.ma_phan_loai_08,
                client.mst, client.dia_chi
            ])

        # Adjust column widths
        for col_num, col_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 20

        # Create a response with the Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="TB_CLIENT_CATEGORIES.xlsx"'
        workbook.save(response)
        return response

# Update xoa_sua field
class UpdateXoaSuaView(APIView):
    def post(self, request):
        ma_khach_hang = request.data.get("ma_khach_hang")
        pass_field = request.data.get("pass_field")
        if not pass_field or pass_field != "admincome":
            return Response({"error": "Invalid or missing password."}, status=status.HTTP_403_FORBIDDEN)

        if not ma_khach_hang:
            return Response({"error": "ma_khach_hang is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if the record exists with xoa_sua = "new"
        record = TB_CLIENT_CATEGORIES.objects.filter(ma_khach_hang=ma_khach_hang, xoa_sua="new").first()
        if not record:
            return Response({"error": "Record with ma_khach_hang does not exist or xoa_sua is not 'new'."}, status=status.HTTP_404_NOT_FOUND)

        # Check the time difference
        time_difference = now() - record.date
        if time_difference < timedelta(hours=0, minutes=2):
            # Update xoa_sua to "delete"
            record.xoa_sua = "delete"
            record.save()
            return Response({"message": "Record updated successfully."}, status=status.HTTP_200_OK)
        else:
            return Response({"error": "overtime to delete"}, status=status.HTTP_400_BAD_REQUEST)
